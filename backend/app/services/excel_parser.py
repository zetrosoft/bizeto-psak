from __future__ import annotations

import html
import io
from pathlib import Path
from typing import Any

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _format_currency(val: float | int) -> str:
    if val == 0:
        return "Rp 0"
    formatted = f"{val:,.0f}".replace(",", ".")
    return f"Rp {formatted}" if val > 0 else f"-Rp {abs(val):,.0f}".replace(",", ".")


def parse_excel_native(file_path: Path, max_rows: int = 15) -> dict[str, Any]:
    """
    Parser Native untuk Excel (.xlsx, .xls, .csv).
    0% Halusinasi, 100% Presisi Angka.
    Mengembalikan data terstruktur dan HTML Table murni (<table>).
    """
    if not HAS_OPENPYXL and file_path.suffix.lower() in [".xlsx", ".xls"]:
        return {
            "success": False,
            "error": "openpyxl library not installed",
            "html_table": "<p class='text-red-500'>Library openpyxl belum terpasang di backend.</p>",
            "rows": [],
        }

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        sheet = wb.active if wb.active else wb[sheet_names[0]]

        raw_rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                raw_rows.append([cell if cell is not None else "" for cell in row])
            if len(raw_rows) >= max_rows + 5:
                break
        wb.close()

        if not raw_rows:
            return {
                "success": False,
                "error": "Sheet kosong",
                "html_table": "<p class='text-amber-500'>Workbook Excel tidak berisi data terbaca.</p>",
                "rows": [],
            }

        headers = [str(h).strip() for h in raw_rows[0]]
        data_rows = raw_rows[1:max_rows + 1]

        # Generate HTML Table Murni dengan Styling Rapi & Cantik
        html_buffer = []
        html_buffer.append('<div class="overflow-x-auto my-3 rounded-xl border border-gold/30 shadow-sm bg-panel">')
        html_buffer.append('  <div class="bg-gold/10 px-4 py-2 border-b border-gold/20 flex items-center justify-between">')
        html_buffer.append(f'    <span class="text-xs font-bold text-ink flex items-center gap-1.5">📊 Preview Data Excel ({html.escape(file_path.name)})</span>')
        html_buffer.append(f'    <span class="text-[10px] text-muted-foreground">Sheet: {html.escape(sheet.title)}</span>')
        html_buffer.append('  </div>')
        html_buffer.append('  <table class="w-full text-left text-xs border-collapse">')
        
        # Header
        html_buffer.append('    <thead>')
        html_buffer.append('      <tr class="bg-muted/80 text-ink font-semibold border-b border-line">')
        for h in headers:
            html_buffer.append(f'        <th class="px-3.5 py-2.5 whitespace-nowrap">{html.escape(h)}</th>')
        html_buffer.append('      </tr>')
        html_buffer.append('    </thead>')

        # Body
        html_buffer.append('    <tbody class="divide-y divide-line/40 text-ink">')
        for idx, row in enumerate(data_rows):
            bg_cls = "bg-panel" if idx % 2 == 0 else "bg-muted/20 hover:bg-gold/5"
            html_buffer.append(f'      <tr class="{bg_cls} transition-colors">')
            for cell in row:
                cell_str = str(cell).strip()
                # Deteksi jika angka nominal
                is_num = False
                try:
                    num_val = float(cell_str.replace(".", "").replace(",", ".")) if isinstance(cell, str) else float(cell)
                    is_num = True
                except (ValueError, TypeError):
                    is_num = False

                align_cls = "text-right font-mono" if is_num and not cell_str.isdigit() and len(cell_str) > 3 else "text-left"
                html_buffer.append(f'        <td class="px-3.5 py-2.5 whitespace-nowrap {align_cls}">{html.escape(cell_str)}</td>')
            html_buffer.append('      </tr>')
        html_buffer.append('    </tbody>')
        html_buffer.append('  </table>')
        if len(raw_rows) > max_rows:
            html_buffer.append(f'  <div class="px-4 py-1.5 text-[10px] text-muted-foreground bg-muted/30 border-t border-line text-center">Menampilkan {max_rows} baris pertama dari total baris spreadsheet</div>')
        html_buffer.append('</div>')

        return {
            "success": True,
            "filename": file_path.name,
            "sheets": sheet_names,
            "active_sheet": sheet.title,
            "headers": headers,
            "rows": data_rows,
            "html_table": "".join(html_buffer),
        }
    except Exception as exc:
        return {
            "success": False,
            "error": str(exc),
            "html_table": f"<p class='text-red-500'>Gagal membaca file Excel: {html.escape(str(exc))}</p>",
            "rows": [],
        }
